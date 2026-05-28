#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook

path = Path('키움_REST_API_문서.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
for ws in wb.worksheets:
    title = ws.title
    if any(k in title for k in ['일봉','차트','ka10081','분봉']):
        print('SHEET', title)
        for i,row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
            vals=[str(x) if x is not None else '' for x in row[:12]]
            line=' | '.join(vals).strip(' |')
            if line:
                print(f'{i}: {line}')
        print('---')
